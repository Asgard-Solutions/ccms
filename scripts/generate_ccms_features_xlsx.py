"""Generate the CCMS feature inventory as an .xlsx matching the
layout of the provided ChiroTouch ChiroPro comparison sheet
(Category | Feature Detail(s) | Status). Drops an extra column for
status so the comparison deck reads at-a-glance.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


OUTPUT = Path("/app/memory/CCMS_Features.xlsx")


# Category -> list of (feature, status) tuples. Status: "shipped",
# "partial", "planned".
FEATURES: list[tuple[str, list[tuple[str, str]]]] = [
    ("Core Platform & Architecture", [
        ("Cloud-based access (preview + prod deploy targets)", "shipped"),
        ("All-in-one platform (EHR + billing + scheduling + payments + communication)", "shipped"),
        ("Multi-provider / multi-location support", "shipped"),
        ("Role-based dashboards and permissions (admin / doctor / staff / patient)", "shipped"),
        ("Multi-tenant data isolation (tenant_id scoping on every collection)", "shipped"),
        ("Data storage, backups, encryption at rest (AES-256-GCM PHI field encryption)", "shipped"),
        ("HIPAA-compliant system (technical safeguards per 45 CFR §164.312)", "shipped"),
        ("Integrated workflows (Clinical ↔ Billing ↔ Scheduling event bus)", "shipped"),
        ("CCPA compliant (data export + soft-delete + retention)", "shipped"),
        ("ISO 27001 alignment (policy + evidence + risk register)", "partial"),
        ("SOC2 certifiable controls (audit log, access review, incident register)", "partial"),
        ("Microservices / event-driven architecture", "shipped"),
        ("Seven-year retention policy with soft-delete on PHI", "shipped"),
    ]),
    ("AI & Automation", [
        ("AI-powered SOAP note generation", "planned"),
        ("AI scribe (voice-to-note)", "planned"),
        ("Context-aware documentation (pulls from prior encounters)", "planned"),
        ("Smart macros auto-population (templates + quick-insert)", "shipped"),
        ("Intake-to-note automation (intake form → clinical_history copy)", "shipped"),
        ("Natural-language scheduling", "planned"),
        ("Automated compliance checks (scrubber + billing readiness + integrity verifier)", "shipped"),
        ("Automated claim scrubbing (17 rules, NPI/dx/modifier/WC checks)", "shipped"),
        ("Auto-emit standalone outcome entries on Re-Exam sign", "shipped"),
        ("Auto follow-up rebook suggestions", "shipped"),
    ]),
    ("EHR & Clinical Documentation", [
        ("SOAP notes (initial exams, follow-up notes, re-exams)", "shipped"),
        ("Customizable macros / templates", "shipped"),
        ("Charting and exam documentation (objective findings, ROM, neuro)", "shipped"),
        ("Patient progression summaries (outcomes trends, pain delta, goals met)", "shipped"),
        ("Clinical templates (chief complaint / HPI / ROS / OMT kinds)", "shipped"),
        ("Audit-ready documentation (signed immutability + append-only addenda)", "shipped"),
        ("CPT code bundling support (charge capture + claim line KIND_TO_HINT mapping)", "shipped"),
        ("Treatment plans with goals, baselines, interventions, discharge criteria", "shipped"),
        ("ICD-10 diagnoses with episode linkage, primary/secondary, activated/resolved", "shipped"),
        ("Imaging & clinical media upload (x-ray, MRI, PDF, 25 MB cap, immutable binary)", "shipped"),
        ("Outcome measures (NDI, Oswestry, Pain VAS, functional index, custom) + SVG trends", "shipped"),
        ("Care timeline v2 (encounters + exams + notes + re-exams + plans + media + outcomes + dx changes)", "shipped"),
        ("Billing readiness evaluator (13 checks, read-only, blocking/warning/ready)", "shipped"),
        ("Append-only signed addenda (reason + narrative, individually signed)", "shipped"),
        ("Clinical audit events (patient-scoped projection + global audit stream)", "shipped"),
        ("Episode / case management (injury / MVA / WC / pediatric / self-pay / insurance)", "shipped"),
    ]),
    ("Patient Intake & Engagement", [
        ("Digital intake forms (chief complaint, HPI, red-flag screening, occupation)", "shipped"),
        ("Patient self check-in", "partial"),
        ("Online booking (patient portal)", "partial"),
        ("Questionnaires (outcome measures delivered to patient)", "partial"),
        ("Two-way texting (Twilio SMS integration)", "partial"),
        ("Appointment reminders (24h + same-day + post-visit review request)", "shipped"),
        ("Review management (review.request notifications, 1-day post-visit)", "shipped"),
        ("Patient communication log (masked by default, unmask with audit)", "shipped"),
        ("Case-specific intake (MVA accident report, WC claim details)", "shipped"),
        ("Break-glass PHI access with reason capture", "shipped"),
    ]),
    ("Scheduling & Patient Flow", [
        ("Smart scheduling (9 visit types, duration defaults, follow-up days)", "shipped"),
        ("Calendar management (day / week views, room assignments)", "shipped"),
        ("Patient flow tracking (scheduled → checked-in → roomed → encounter → checkout)", "shipped"),
        ("Daily schedule snapshot (today's appointments on dashboard)", "shipped"),
        ("Flow board (queue view per location)", "shipped"),
        ("Appointment workflow panel (check-in / no-show / launch encounter / reschedule)", "shipped"),
        ("Room management (exam / consult / x-ray / therapy types)", "shipped"),
        ("Appointment type catalog (sort order, color, default duration)", "shipped"),
        ("Provider filter / My Queue", "shipped"),
        ("Checkout queue with follow-up rebook", "shipped"),
    ]),
    ("Payments & Financial", [
        ("Integrated payment processing (Stripe playbook ready to wire)", "planned"),
        ("Recurring / saved payment methods", "planned"),
        ("Ledgers (patient ledger card + page with allocations)", "shipped"),
        ("Payment reconciliation (post-payment dialog, invoice allocation)", "shipped"),
        ("Invoice management (detail, list, statement generation)", "shipped"),
        ("Patient statements (balance, invoice count, as-of-date)", "shipped"),
        ("A/R aging report", "shipped"),
        ("Multi-currency fields (USD default, legacy-tolerant)", "shipped"),
        ("Charge capture dialog", "shipped"),
        ("Refund / void tracking", "partial"),
    ]),
    ("Billing & RCM", [
        ("Electronic claim submission (837P wire + JSON payload generation)", "shipped"),
        ("Claim scrubbing (17 rules, NPI, modifier, subluxation-dx, WC case#)", "shipped"),
        ("Eligibility verification (270/271)", "planned"),
        ("ERA auto-posting (835 import + remittance posting)", "shipped"),
        ("Denial prevention (billing readiness + scrubber)", "shipped"),
        ("Medicare / PI / WC support (payer catalog, case-type routing, portal vs EDI)", "shipped"),
        ("Claims queue (14-row multi-lifecycle, 4 tabs: All / Pending / Denied / Follow-up)", "shipped"),
        ("Claim assignment (tenant-scoped assignee dropdown, Assign to me, unassign)", "shipped"),
        ("Follow-up flagging with reason + next-action date", "shipped"),
        ("Submission lifecycle (draft → ready → submitted → paid / denied / rejected / partial)", "shipped"),
        ("837P payload preview (ISA*GS*ST envelope, NM1 name-split, decrypted addresses)", "shipped"),
        ("Human-readable refs (no UUID leakage; real NPIs, MRNs, member IDs)", "shipped"),
        ("Portal-submitted claim support (PIP, WC; correctly no 837P)", "shipped"),
        ("Payer catalog + electronic payer IDs", "shipped"),
        ("Fee schedules (per-payer, effective dates)", "shipped"),
        ("Remittance import (raw 835) + posting", "shipped"),
        ("Denials queue (denial codes, rebill story)", "shipped"),
        ("Claim-from-encounter synthesis (readiness-gated, auto dx + CPT hints)", "shipped"),
        ("Claim timeline (history + scrubber runs + submissions + outcomes)", "shipped"),
    ]),
    ("Compliance", [
        ("Real-time compliance scan (billing readiness evaluator per encounter)", "shipped"),
        ("Coding validation (ICD-10 + CPT + modifier + dx linkage)", "shipped"),
        ("Risk flagging (13-check billing readiness + scrubber blocking rules)", "shipped"),
        ("HIPAA controls (masked-by-default PHI, break-glass, reauth, soft-delete, audit)", "shipped"),
        ("AES-256-GCM PHI field encryption at rest", "shipped"),
        ("Audit log viewer (admin-only, full action trail)", "shipped"),
        ("Clinical audit events (patient-scoped; filterable by event type)", "shipped"),
        ("Compliance Ops surface (Evidence / Incidents / Vendors / Risks / Policies / Access Reviews)", "shipped"),
        ("Access review cycles", "partial"),
        ("Integrity verifier (foreign-key checks, duplicate detection, shape locks)", "shipped"),
        ("Immutable signed notes + append-only addenda", "shipped"),
        ("Session idle timeout (15 min) + suspicious-login step-up", "shipped"),
    ]),
    ("Reporting & Analytics", [
        ("Dashboards (billing, clinical summary, today's schedule)", "shipped"),
        ("Revenue tracking (Outstanding Balance, Lifetime Billed, Payments Recorded)", "shipped"),
        ("Payer analysis / payer mix", "partial"),
        ("Custom reports (reports landing page + report viewer)", "partial"),
        ("A/R aging", "shipped"),
        ("Denial heat map", "partial"),
        ("Outcome trends per measure (inline SVG charts)", "shipped"),
        ("Clinical summary counts (visits / exams / plans / re-exams / notes / dx)", "shipped"),
        ("Patient progression tracking (baseline → current across re-exams)", "shipped"),
    ]),
    ("Integrations", [
        ("Clearinghouse integrations (Change / Optum sandbox ready)", "partial"),
        ("Payment integrations (Stripe playbook)", "partial"),
        ("Helcim", "planned"),
        ("API ecosystem (FastAPI /api gateway, JWT auth, RBAC, audit)", "shipped"),
        ("Twilio SMS", "partial"),
        ("Resend email", "partial"),
        ("Emergent object storage (clinical media, upload/download/soft-delete)", "shipped"),
        ("Emergent LLM key (OpenAI / Claude / Gemini text, Nano Banana, Whisper)", "shipped"),
        ("Emergent-managed Google OAuth", "planned"),
    ]),
    ("Security & Identity (CCMS-specific)", [
        ("Bcrypt password hashing + JWT cookie sessions", "shipped"),
        ("TOTP MFA with 8 backup codes", "shipped"),
        ("6-digit PIN (masked, lockout, rotation)", "shipped"),
        ("Step-up reauth (password OR PIN, 5-min reauth token)", "shipped"),
        ("Password policy (12-char complexity, history reuse rejection)", "shipped"),
        ("Change-password hardening (rate limit 5 fails/15 min per user, 60/min per IP)", "shipped"),
        ("Password reset (single-use 15-min tokens)", "shipped"),
        ("Per-user / per-IP brute-force protection", "shipped"),
        ("Session epoch bump on email change (invalidates stale tokens)", "shipped"),
        ("Role and permission overrides (admin-managed, revocable, audited)", "shipped"),
        ("Workforce management (onboarding, role assignment, enable/disable)", "shipped"),
        ("Elevation requests (temporary admin rights with audit)", "shipped"),
        ("Recent sign-ins viewer per user", "shipped"),
        ("PHI masking toggle with audit on unmask", "shipped"),
        ("Data export (patient self-service + admin)", "shipped"),
    ]),
]


STATUS_FILL = {
    "shipped": PatternFill("solid", fgColor="D4F4DD"),
    "partial": PatternFill("solid", fgColor="FFF3C4"),
    "planned": PatternFill("solid", fgColor="F8D7DA"),
}
STATUS_LABEL = {
    "shipped": "✅ Shipped",
    "partial": "🟡 Partial",
    "planned": "⛔ Planned",
}
HEADER_FILL = PatternFill("solid", fgColor="1F3B4D")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
CAT_FILL = PatternFill("solid", fgColor="E6EEF3")
CAT_FONT = Font(bold=True, size=11, name="Calibri", color="1F3B4D")
BODY_FONT = Font(size=10, name="Calibri")
THIN = Side(border_style="thin", color="C4CCD2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_sheet() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "CCMS Feature Inventory"

    # Header row
    headers = ["CCMS Category", "Feature Detail(s)", "Status"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 22

    row = 2
    total = 0
    shipped = partial = planned = 0
    for category, items in FEATURES:
        for feature, status in items:
            c_cat = ws.cell(row=row, column=1, value=category)
            c_feat = ws.cell(row=row, column=2, value=feature)
            c_stat = ws.cell(row=row, column=3, value=STATUS_LABEL[status])
            c_cat.fill = CAT_FILL
            c_cat.font = CAT_FONT
            c_feat.font = BODY_FONT
            c_stat.font = BODY_FONT
            c_stat.fill = STATUS_FILL[status]
            for c in (c_cat, c_feat, c_stat):
                c.border = BORDER
                c.alignment = Alignment(vertical="center", wrap_text=True)
            row += 1
            total += 1
            if status == "shipped":
                shipped += 1
            elif status == "partial":
                partial += 1
            else:
                planned += 1

    # Column widths
    widths = {"A": 36, "B": 90, "C": 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    # Summary sheet
    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Count"])
    for col in (1, 2):
        cell = summary.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
    summary.append(["Total features catalogued", total])
    summary.append(["✅ Shipped", shipped])
    summary.append(["🟡 Partial", partial])
    summary.append(["⛔ Planned", planned])
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 12
    for r in range(2, 6):
        for col in (1, 2):
            summary.cell(row=r, column=col).border = BORDER
            summary.cell(row=r, column=col).font = BODY_FONT

    # Per-category summary
    summary.append([])
    summary.append(["Category", "Features"])
    for col in (1, 2):
        cell = summary.cell(row=summary.max_row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
    for category, items in FEATURES:
        summary.append([category, len(items)])
        for col in (1, 2):
            summary.cell(row=summary.max_row, column=col).border = BORDER
            summary.cell(row=summary.max_row, column=col).font = BODY_FONT

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT} — {total} features across {len(FEATURES)} categories")
    print(f"  Shipped: {shipped} · Partial: {partial} · Planned: {planned}")


if __name__ == "__main__":
    write_sheet()
