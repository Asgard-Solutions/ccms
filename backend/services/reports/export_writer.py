"""
Report export writers — CSV, Excel (xlsx), and PDF.

Security model (HIPAA):
  * **PDF** — When a password is supplied we encrypt the PDF *natively*
    using reportlab's StandardEncryption (AES-128, PDF v4 header). Readers
    prompt for the password when opened; no wrapper is used.
  * **CSV and XLSX** — These formats cannot be truly encrypted with an
    open-source library without pulling in heavyweight office tooling
    (openpyxl's ``workbook.security`` is a *protection flag* only, not
    encryption). Instead we wrap the file in an AES-256 password-protected
    ZIP archive (pyzipper) which *is* real encryption.
  * A file is only ever labelled ``password_protected=True`` when its
    bytes on disk genuinely require the password to read.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pyzipper
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors, pdfencrypt
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from services.reports.definitions import Column


# ---------------------------------------------------------------------------
# Value formatting for human-facing output
# ---------------------------------------------------------------------------

def _format_value(val: Any, col: Column) -> Any:
    if val is None or val == "":
        return ""
    if col.type == "currency":
        try:
            cents = int(val)
        except (TypeError, ValueError):
            return val
        sign = "-" if cents < 0 else ""
        whole, fraction = divmod(abs(cents), 100)
        return f"{sign}${whole:,}.{fraction:02d}"
    if col.type == "boolean":
        return "Yes" if val else "No"
    if col.type == "datetime" and isinstance(val, str) and len(val) >= 10:
        return val.replace("T", " ")[:19]
    return val


def _rows_for_output(rows: list[dict[str, Any]], columns: list[Column]) -> list[list[Any]]:
    return [[_format_value(r.get(c.key), c) for c in columns] for r in rows]


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def _write_csv(path: Path, columns: list[Column], rows: list[dict[str, Any]]) -> None:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([c.label for c in columns])
    for r in _rows_for_output(rows, columns):
        w.writerow(r)
    path.write_text(buf.getvalue(), encoding="utf-8")


# ---------------------------------------------------------------------------
# Excel (.xlsx)
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")


def _write_xlsx(path: Path, columns: list[Column], rows: list[dict[str, Any]], title: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Report")[:31]

    # Title row
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(columns)))

    # Header row
    header_row = 3
    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=idx, value=col.label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Data rows
    for r_idx, r in enumerate(_rows_for_output(rows, columns), start=header_row + 1):
        for c_idx, val in enumerate(r, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            col = columns[c_idx - 1]
            if col.type in ("number", "integer", "currency"):
                cell.alignment = Alignment(horizontal="right")

    # Column widths — cheap auto-sizing
    for idx, col in enumerate(columns, start=1):
        letter = ws.cell(row=header_row, column=idx).column_letter
        ws.column_dimensions[letter].width = max(14, min(40, len(col.label) + 6))

    ws.freeze_panes = f"A{header_row + 1}"
    wb.save(path)


# ---------------------------------------------------------------------------
# PDF (reportlab)
# ---------------------------------------------------------------------------

def _write_pdf(
    path: Path,
    columns: list[Column],
    rows: list[dict[str, Any]],
    title: str,
    *,
    password: str | None = None,
) -> None:
    """Render the PDF, optionally encrypting it natively with AES-128.

    When `password` is supplied the resulting PDF opens in any standard
    reader with that password (no ZIP wrapper). We never print, or copy
    blocked — the password is purely an access control gate.
    """
    encrypt = None
    if password:
        encrypt = pdfencrypt.StandardEncryption(
            userPassword=password,
            ownerPassword=password,
            canPrint=1, canModify=0, canCopy=1, canAnnotate=0,
            strength=128,
        )
    doc = SimpleDocTemplate(
        str(path),
        pagesize=landscape(letter),
        leftMargin=0.4 * inch, rightMargin=0.4 * inch,
        topMargin=0.5 * inch, bottomMargin=0.5 * inch,
        title=title,
        encrypt=encrypt,
    )
    styles = getSampleStyleSheet()
    story: list = []
    story.append(Paragraph(title, styles["Title"]))
    story.append(Paragraph(
        f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC · {len(rows)} rows",
        styles["Normal"],
    ))
    story.append(Spacer(1, 8))

    data = [[c.label for c in columns]] + _rows_for_output(rows, columns)
    # Trim very long cells for PDF readability
    data = [[(str(v)[:80] + "…") if isinstance(v, str) and len(v) > 80 else v for v in row] for row in data]

    tbl = Table(data, repeatRows=1)
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("ALIGN", (0, 0), (-1, 0), "LEFT"),
    ])
    tbl.setStyle(style)
    story.append(tbl)
    doc.build(story)


# ---------------------------------------------------------------------------
# Password-protected ZIP wrapper (AES-256)
# ---------------------------------------------------------------------------

def _zip_with_password(src_path: Path, zip_path: Path, password: str) -> None:
    pwd = password.encode("utf-8")
    with pyzipper.AESZipFile(
        zip_path, "w",
        compression=pyzipper.ZIP_DEFLATED,
        encryption=pyzipper.WZ_AES,
    ) as zf:
        zf.setpassword(pwd)
        zf.setencryption(pyzipper.WZ_AES, nbits=256)
        zf.write(src_path, arcname=src_path.name)
    try:
        src_path.unlink()
    except OSError:
        pass


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

@dataclass
class ExportArtifact:
    path: Path
    mime: str
    filename: str
    password_protected: bool
    size_bytes: int
    protection_kind: str  # "none" | "pdf_native" | "aes_zip"


MIME_TYPES = {
    "csv": "text/csv",
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
    "zip": "application/zip",
}


_SLUG_STRIP = re.compile(r"[^A-Za-z0-9._-]+")


def _slugify(title: str) -> str:
    """Whitespace → `_`, non-alnum → `-`, collapse repeats, max 60 chars."""
    slug = title.strip().replace(" ", "_")
    slug = _SLUG_STRIP.sub("-", slug).strip("-_")
    return slug[:60] or "report"


def _human_filename(title: str, ext: str, *, when: datetime | None = None) -> str:
    when = when or datetime.now(timezone.utc)
    return f"{_slugify(title)}-{when.strftime('%Y%m%d-%H%M')}.{ext}"


def build_export(
    *,
    dest_dir: Path,
    export_id: str,
    title: str,
    columns: list[Column],
    rows: list[dict[str, Any]],
    fmt: str,
    password: str | None = None,
) -> ExportArtifact:
    """Build the requested format at `dest_dir/{export_id}.{ext}`.

    Protection rules:
      * `fmt="pdf"` + password → native PDF encryption (AES-128). No ZIP.
      * `fmt="csv"`/`"excel"` + password → AES-256 password-protected ZIP
        (pyzipper). CSV and native xlsx cannot be encrypted without the
        archive wrapper.
      * No password → plain file.
    """
    dest_dir.mkdir(parents=True, exist_ok=True)
    raw_ext = {"csv": "csv", "excel": "xlsx", "pdf": "pdf"}[fmt]
    raw_path = dest_dir / f"{export_id}.{raw_ext}"

    protection = "none"

    if fmt == "csv":
        _write_csv(raw_path, columns, rows)
    elif fmt == "excel":
        _write_xlsx(raw_path, columns, rows, title)
    elif fmt == "pdf":
        _write_pdf(raw_path, columns, rows, title, password=password)
        if password:
            protection = "pdf_native"
    else:  # pragma: no cover
        raise ValueError(f"Unsupported format: {fmt}")

    # Wrap CSV/XLSX in an encrypted ZIP when a password is required.
    if password and fmt in ("csv", "excel"):
        zip_path = dest_dir / f"{export_id}.zip"
        _zip_with_password(raw_path, zip_path, password)
        protection = "aes_zip"
        return ExportArtifact(
            path=zip_path,
            mime=MIME_TYPES["zip"],
            filename=_human_filename(title, "zip"),
            password_protected=True,
            size_bytes=zip_path.stat().st_size,
            protection_kind=protection,
        )

    # Native-encrypted PDF keeps its .pdf extension.
    return ExportArtifact(
        path=raw_path,
        mime=MIME_TYPES[fmt],
        filename=_human_filename(title, raw_ext),
        password_protected=protection != "none",
        size_bytes=raw_path.stat().st_size,
        protection_kind=protection,
    )


def generate_password(length: int = 20) -> str:
    """URL-safe one-time password. 20 chars ≈ 120 bits of entropy."""
    import secrets
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghjkmnpqrstuvwxyz23456789"
    return "".join(secrets.choice(alphabet) for _ in range(length))
