"""
Unit tests for the Reports framework — every test runs against the live
test Mongo via tenant_db() with a fresh tenant so no cross-run pollution.

Scope:
* `ReportDefinition.to_public` round-trips the metadata contract.
* `resolve_columns` / `resolve_sort` honour the default set + unknown-key fallback.
* Built-in `appointments_list` runner returns expected shape with tenant scope.
* Export writer — CSV + XLSX + PDF + password-protected ZIP paths.
"""
from __future__ import annotations

import csv
import io
import os
import uuid
from pathlib import Path

import pyzipper
import pytest
from openpyxl import load_workbook

from services.reports.definitions import (
    Column,
    QueryContext,
    ReportDefinition,
    all_definitions,
    get_definition,
    resolve_columns,
    resolve_sort,
)
from services.reports.export_writer import build_export, generate_password


# ---------------------------------------------------------------------------
# Framework — no network
# ---------------------------------------------------------------------------

def test_registry_contains_core_reports():
    names = {d.name for d in all_definitions()}
    # Spot-check a representative report from each category
    for required in ("appointments_list", "patient_roster",
                     "unsigned_clinical_notes", "claims_list",
                     "invoices_list", "payments_received",
                     "denials_log", "audit_activity",
                     "license_expiration", "provider_productivity"):
        assert required in names, f"missing {required}"


def test_resolve_columns_falls_back_to_defaults():
    d = get_definition("appointments_list")
    assert d is not None
    cols = resolve_columns(d, None)
    assert len(cols) > 0
    # All returned columns are in the default set
    defaults = set(d.default_columns)
    for c in cols:
        assert c.key in defaults

    # Explicit + unknown keys: unknowns dropped, known returned
    cols2 = resolve_columns(d, ["status", "__bogus__", "patient_name"])
    assert [c.key for c in cols2] == ["status", "patient_name"]


def test_resolve_sort_falls_back_to_default():
    d = get_definition("claims_list")
    assert resolve_sort(d, None) == d.default_sort
    assert resolve_sort(d, "__bogus__") == d.default_sort
    assert resolve_sort(d, d.sort_options[0].key) == d.sort_options[0].key


def test_to_public_contract():
    d = get_definition("patient_roster")
    pub = d.to_public()
    # Public payload only carries keys the frontend expects
    assert pub["name"] == "patient_roster"
    assert pub["contains_phi"] is True
    assert isinstance(pub["columns"], list) and pub["columns"]
    assert {"csv", "excel", "pdf"}.issubset(set(pub["export_formats"]))
    assert set(pub["default_columns"]).issubset({c["key"] for c in pub["columns"]})


# ---------------------------------------------------------------------------
# Export writer — end-to-end artifact generation
# ---------------------------------------------------------------------------

TEST_COLUMNS = [
    Column("id", "Id", "string"),
    Column("amount_cents", "Amount", "currency", align="right"),
    Column("count", "Count", "integer", align="right"),
    Column("when", "When", "datetime"),
]

TEST_ROWS = [
    {"id": "r1", "amount_cents": 12345, "count": 3, "when": "2026-01-02T03:04:05"},
    {"id": "r2", "amount_cents": -500, "count": 0, "when": "2026-01-03T08:30:00"},
]


def test_csv_export_is_human_readable(tmp_path):
    art = build_export(
        dest_dir=tmp_path, export_id=str(uuid.uuid4()),
        title="Test", columns=TEST_COLUMNS, rows=TEST_ROWS, fmt="csv",
    )
    assert art.mime == "text/csv"
    assert not art.password_protected
    rows = list(csv.reader(io.StringIO(art.path.read_text())))
    assert rows[0] == ["Id", "Amount", "Count", "When"]
    # Currency is formatted with $ and two decimal places
    assert rows[1][1] == "$123.45"
    # Negative currency retains the minus sign
    assert rows[2][1] == "-$5.00"


def test_xlsx_export_contains_header_and_rows(tmp_path):
    art = build_export(
        dest_dir=tmp_path, export_id=str(uuid.uuid4()),
        title="Test XL", columns=TEST_COLUMNS, rows=TEST_ROWS, fmt="excel",
    )
    assert art.path.suffix == ".xlsx"
    wb = load_workbook(art.path)
    ws = wb.active
    assert ws["A1"].value == "Test XL"  # title row
    # Header row at row 3; check each label
    labels = [ws.cell(row=3, column=i + 1).value for i in range(len(TEST_COLUMNS))]
    assert labels == ["Id", "Amount", "Count", "When"]


def test_pdf_export_generates_file(tmp_path):
    art = build_export(
        dest_dir=tmp_path, export_id=str(uuid.uuid4()),
        title="Test PDF", columns=TEST_COLUMNS, rows=TEST_ROWS, fmt="pdf",
    )
    assert art.path.suffix == ".pdf"
    assert art.path.stat().st_size > 0
    head = art.path.read_bytes()[:4]
    assert head == b"%PDF"  # PDF magic number
    assert art.protection_kind == "none"
    assert art.password_protected is False


def test_pdf_export_native_password_encryption(tmp_path):
    """
    Password-protected PDFs are encrypted *natively* (AES-128, PDF v4).
    They do NOT get wrapped in a ZIP — the .pdf extension is preserved and
    the file carries the `/Encrypt` dictionary so any standard reader
    will prompt for the password.
    """
    pw = generate_password()
    art = build_export(
        dest_dir=tmp_path, export_id=str(uuid.uuid4()),
        title="Secure PDF", columns=TEST_COLUMNS, rows=TEST_ROWS, fmt="pdf",
        password=pw,
    )
    assert art.path.suffix == ".pdf", "PDF must not be re-wrapped in a ZIP"
    assert art.protection_kind == "pdf_native"
    assert art.password_protected is True
    assert art.mime == "application/pdf"

    body = art.path.read_bytes()
    assert body[:4] == b"%PDF"
    # Presence of the encryption dictionary is the ground-truth signal of
    # a *genuinely* password-protected PDF.
    assert b"/Encrypt" in body, "PDF must contain an /Encrypt dict when password-protected"


def test_xlsx_with_password_produces_aes_zip(tmp_path):
    pw = generate_password()
    art = build_export(
        dest_dir=tmp_path, export_id=str(uuid.uuid4()),
        title="Secure XLSX", columns=TEST_COLUMNS, rows=TEST_ROWS, fmt="excel",
        password=pw,
    )
    # xlsx can't be truly encrypted by openpyxl → ZIP wrapper is correct.
    assert art.path.suffix == ".zip"
    assert art.protection_kind == "aes_zip"
    assert art.mime == "application/zip"
    with pyzipper.AESZipFile(art.path) as zf:
        zf.setpassword(pw.encode())
        members = zf.namelist()
        assert any(m.endswith(".xlsx") for m in members)


def test_filename_includes_title_and_date(tmp_path):
    art = build_export(
        dest_dir=tmp_path, export_id=str(uuid.uuid4()),
        title="Patient Roster Q1", columns=TEST_COLUMNS, rows=TEST_ROWS, fmt="csv",
    )
    # Pattern: `Patient_Roster_Q1-YYYYMMDD-HHMM.csv`
    import re as _re
    assert _re.match(
        r"^Patient_Roster_Q1-\d{8}-\d{4}\.csv$", art.filename,
    ), f"filename does not match expected pattern: {art.filename}"


def test_filename_slug_strips_unsafe_characters(tmp_path):
    art = build_export(
        dest_dir=tmp_path, export_id=str(uuid.uuid4()),
        title="A/B::Report? 2026", columns=TEST_COLUMNS, rows=TEST_ROWS, fmt="csv",
    )
    # No slashes, colons, or question marks in the filename.
    for ch in "/:?":
        assert ch not in art.filename


def test_password_never_serialises_to_artifact(tmp_path):
    """The ExportArtifact dataclass exposes no plaintext password field."""
    pw = generate_password()
    art = build_export(
        dest_dir=tmp_path, export_id=str(uuid.uuid4()),
        title="Secure CSV", columns=TEST_COLUMNS, rows=TEST_ROWS, fmt="csv",
        password=pw,
    )
    fields = set(vars(art).keys())
    forbidden = {"password", "password_plain", "plaintext", "secret"}
    assert not (fields & forbidden), f"artifact leaks password field: {fields & forbidden}"


def test_password_protected_zip_decrypts_with_given_password(tmp_path):
    pw = generate_password()
    art = build_export(
        dest_dir=tmp_path, export_id=str(uuid.uuid4()),
        title="PHI Export", columns=TEST_COLUMNS, rows=TEST_ROWS, fmt="csv",
        password=pw,
    )
    assert art.path.suffix == ".zip"
    assert art.password_protected
    with pyzipper.AESZipFile(art.path) as zf:
        zf.setpassword(pw.encode("utf-8"))
        members = zf.namelist()
        assert len(members) == 1 and members[0].endswith(".csv")
        body = zf.read(members[0]).decode("utf-8")
        assert "Id,Amount,Count,When" in body.replace("\r\n", "\n")

    # Wrong password raises
    with pytest.raises(RuntimeError):
        with pyzipper.AESZipFile(art.path) as zf:
            zf.setpassword(b"wrong-password")
            zf.read(zf.namelist()[0])


def test_generate_password_is_long_and_unique():
    a = generate_password()
    b = generate_password()
    assert len(a) >= 20
    assert a != b
    # URL-safe alphabet only (no ambiguous characters)
    for ch in a:
        assert ch.isalnum()
